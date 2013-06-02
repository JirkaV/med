from openpyxl.reader.excel import load_workbook

def process_xlsx(fileobj):
    wb = load_workbook(fileobj)

    prehled = wb.worksheets[0]
    ws = wb.worksheets[2]

    rc = prehled.cell('B2').value
    prijmeni = prehled.cell('D2').value
    jmeno = prehled.cell('I2').value
#    print rc, prijmeni, jmeno

    data_range = ws.range('A3:CB140')
    data = []
    for row in data_range:
        datum = row[0].value
        if not datum:
            continue
        datum = datum.isoformat()
        viremie = row[71].value
        try:
            viremie = int(viremie)
        except TypeError:
            continue  # None
        except ValueError:
            if viremie.lower().startswith('neg'):
                viremie = 0
            elif 'ed' in viremie and 'na' in viremie:  # seda zona
                viremie = 100
            else:
                continue  # Huh?
        data.append([datum, viremie])
    return {'rc': rc,
            'jmeno': jmeno,
            'prijmeni': prijmeni,
            'data': data}
