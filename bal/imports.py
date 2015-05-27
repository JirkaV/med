# -*- coding: utf-8 -*-
import sys
import os
import datetime
import glob
import django
import re
from openpyxl.reader.excel import load_workbook
from pgs.bal.models import *

re_DG = re.compile('^[A-Z] ?[0-9]{3}')

DG = {
'HL': u'HL',
'DLBCL CNS': u'NHL',
'CLL': u'CLL/CLL',
'?': u'<FIXME>',
'ALL': u'ALL',
'APL': u'AML',
'ALCL': u'NHL',
'HCL': u'ostatni hematologicke malignity',
'ALCL Anaplastic large-cell lymphoma': u'NHL',
'CASTELMAN': u'nemaligni hematologicka onemocneni',
'AML': u'AML',
'AGRANUL': u'<FIXME>',
'DLBCL': u'NHL',
'CML': u'CML',
'MM': u'MM',
'MDS': u'MDS',
'MCL': u'NHL',
'FL': u'NHL',
'AGRANULOC': u'<FIXME>',
'SLL': u'CLL/SLL',
'PTCL': u'NHL',
'ITP': u'nemaligni hematologicka onemocneni',
'T-NHL': u'NHL',
'MPD': u'MPD + MDS/MPD',
'MALT PLIC': u'NHL',
'CMML2': u'MPD + MDS/MPD',
'ANGIOIMUNOBLASTICK+L T-LY': u'NHL',
'ANPLASTICK+L T LYMFOM': u'NHL',
'MYELOM': u'MM',
'CLL, INVAZ. ASPERGILOZA': u'CLL/SLL',
'SAA, amyloid¢za': u'nemaligni hematologicka onemocneni',
'FOLIKUL+ÅRN+Z LYMFOM': u'NHL',
'FOLIKULµRN÷ LYMFOM': u'NHL',
'AL AMYLOIDOZA , amyloid¢za': u'nemaligni hematologicka onemocneni',
'HEMOFAGOCYTOZA, PCP': u'nemaligni hematologicka onemocneni',
'BURKITT': u'NHL',
'AK. BIFENOTYPOV+Å LEUKEMIE': u'ostatni hematologicke malignity',
'CNS LYMFOM': u'NHL',
'HODGKIN+ÆV LYMFOM': u'HL',
'D695': u'nemaligni hematologicka onemocneni',
'VLASATOBUN-ö-åN+Å LEUKEMIE': u'ostatni hematologicke malignity',
'APLASTICK+Å ANEMIE': u'aplasticka anemie',
'APLASTICKµ ANEMIE': u'aplasticka anemie',
'NHL': u'NHL',
'SAA': u'nemaligni hematologicka onemocneni',
'T NHL': u'NHL',
'B-ALL, BCR/ABL +': u'ALL',
'AML-SEK': u'AML',
'PNEUMOKOKOV+Å PNEUMONIE': u'<FIXME>',
'ADENOCA ENDOMETRIA': u'<IGNORE>',
'AML - SEK': u'AML',
'MDS, ZYGOMYKOZA - KOLENO': u'MDS',
'CA PLIC METASTAZUJ+ZC+Z': u'<IGNORE>',
'SLE': u'<IGNORE>',
'CA PRSU': u'<IGNORE>',
'CMML': u'MPD + MDS/MPD',
'SYS ONEM POJIVA': u'<IGNORE>',
'PCL': u'ostatni hematologicke malignity',
'CUSHING': u'<IGNORE>',
'MZL': u'NHL',
'ANGIOIMUNO NHL': u'NHL',
'AGRAN PO THYROZOLU': u'<IGNORE>',
'INCIDENTALOM NADLEDVINY': u'<IGNORE>',
'BL': u'NHL',
'C833': u'NHL',
'C919': u'ALL',
'C833': u'NHL',
'C900': u'MM',
'C920': u'AML',
'C911': u'CLL/SLL',
'D469, NS': u'MDS',
'C840': u'NHL',
'AML Z MDS': u'AML',
'AKUTN+Z PRE-B LYMFOBLASTICK+Å LEUKEMIE': u'ALL',
'C811': u'HL',
'C831': u'NHL',
'C827': u'NHL',
'C911': u'CLL/SLL',
'D619': u'aplasticka anamie',
'C811': u'HL',
'D589': u'nemaligni hematologicka onemocneni',
'C833': u'NHL',
'D469': u'ostatni hematologicke malignity',
'C920': u'AML',
'C914': u'ostatni hematologicke malignity',
'C920': u'AML',
'C837': u'NHL',
'D469': u'MDS',
'C812': u'HL',
'D619': u'aplasticka anamie',
'C920': u'AML',
'D462': u'MDS',
'D467': u'MDS',
'D467': u'MPD + MDS/MPD',
'D613': u'aplasticka anamie',
'C833': u'NHL',
'C910': u'ALL',
'C921': u'ALL',
'C811': u'HL',
'C910': u'ALL',
'D462': u'MDS',
'C927': u'AML',
'C910': u'ALL',
'C900': u'MM',
'C911': u'CLL/SLL',
'D471': u'MPD + MDS/MPD',
'C911': u'CLL/SLL',
'C920': u'AML',
'C931': u'MPD + MDS/MPD',
'D591': u'nemaligni hematologicka onemocneni',
'C845': u'NHL',
'C833': u'NHL',
'D462': u'MDS',
'C924': u'AML',
'C920': u'AML',
'C921': u'CML',
'C820': u'NHL',
'C827': u'NHL',
'C811': u'NL',
'C844': u'NHL',
'BLASTICK+L ZVRAT CHRONICK+â MYELOIDN+Z LEUK+âMIE, 19.8.2010, BLASTY LYMFOIDN+Z, IMUNOFENOTYP CALLA (CD10+, CD34+)': u'ALL',
'C833': u'NHL',
'D693': u'nemaligni hematologicka onemocneni',
'D467': u'MDS',
'C901': u'MM',
'C 901': u'MM',
'C959': u'ostatni hematologicke malignity',
'C829': u'NHL',
'C830': u'NHL',
'C812': u'HL',
'C851': u'NHL',
'C813': u'HL',
}


# def process_line(entries):
#     rc = entries[1]
#     if not '/' in rc:
#       return # hlavicka ?
#     pacient, _ = Pacient.objects.get_or_create(rc=rc)
#
#     diagnoza = entries[2]
#     if not diagnoza:
#         return
#     kod, popis = diagnoza.split(None, 1)
#     kod = unicode(kod)
#     dg, created = Diagnoza.objects.get_or_create(kod=kod)
#     if created:
#         dg.popis = popis.decode('cp1250').strip()
#         dg.save()
#     pacient.diagnozy.add(dg)
#     pacient.save()
#
#     odber = datetime.datetime.strptime(entries[3].strip(), '%d.%m.%y').date()
#     o, _ = Odber.objects.get_or_create(pacient=pacient, datum=odber)
#
#     for prukaz in Prukaz.objects.all():
#         try:
#             v = entries[prukaz.sloupec - 1].strip()
#         except IndexError:
#             print '[!] Chybny radek (%s)' % rc
#             return
#         if 'neg' in v:
#             vysledek = 'neg'
#         elif ('poz' in v or 'ano' in v):
#             vysledek = 'poz'
#         elif ('ed' in v and ' z' in v):
#             vysledek = 'sz' # seda zona
#         elif 'hran' in v:
#             vysledek = 'hran'
#         elif 'nehod' in v:
#             vysledek = 'nehod'
#         elif (not v or v == 'x'):
#             vysledek = 'x'
#         else:
#             print '[!] Neznamy vysledek mereni: "%s"' % v
#             return
#         vysl, _ = VysledekTestu.objects.get_or_create(odber=o,
#                                                       prukaz=prukaz,
#                                                       vysledek=vysledek)
#
# def learn_prukaz(entries):
#     col = 0
#     for e in entries:
#         e = e.strip()
#         col += 1
#         if e:
#             p, created = Prukaz.objects.get_or_create(sloupec=col)
#             if created:
#                 p.jmeno = e.decode('cp1250')
#                 p.save()

django.setup()  # avoid AppRegistryNotReady

def clean_dg(raw_dg, rc):
    raw_dg = raw_dg.strip()
    dg_no = re_DG.search(raw_dg)
    if dg_no is not None:
        raw_dg = dg_no.group(0)
    try:
        new_dg = unicode(DG[raw_dg])
    except KeyError:
        print 'Missing DG "%s" [%s]' % (raw_dg, rc)
        return None
    res, _ = Diagnoza.objects.get_or_create(kod=new_dg)
    return res

def import_summary(summary_generator):
    for row in summary_generator:
        rc = row[0]
        datum_odberu = row[1]
        dg = row[2]['diagnoza'].upper()
        diagnoza, d_created = Diagnoza.objects.get_or_create(kod=dg)
        p, created = Pacient.objects.get_or_create(rc=rc)
        p.diagnozy.add(diagnoza)
        if p.dg is None:
            cleaned_dg = clean_dg(dg, rc)
            p.dg = cleaned_dg
            p.save()
        bal, _ = BAL.objects.get_or_create(pacient=p,
                                           datum=datum_odberu,
                                           alogenni_transplantace=row[2]['alot'],
                                           agranulocytoza=row[2]['agranulocytoza'],
                                           exitus=row[2]['exitus'])

def import_bal(bal_generator):
    chybejici_pacienti = set()
    chybejici_baly = set()
    for row in bal_generator:
        try:
            if row[0] is None:
                break
            p = Pacient.objects.get(rc=row[0])
            try:
                bal = BAL.objects.get(pacient=p, datum=row[1])
                agens, _ = Agens.objects.get_or_create(jmeno=row[2]['nalez'])
                odber, _ = Odber.objects.get_or_create(bal=bal,
                                                       nalez=agens,
                                                       kultivace_typ=row[2]['pp'],
                                                       kultivace_vysledek=row[2]['vysledek'],
                                                       kultivace_memo=row[2]['memo'])
            except BAL.DoesNotExist:
                chybejici_baly.add( (p.rc, row[1]) )
                continue
        except Pacient.DoesNotExist:
            chybejici_pacienti.add(row[0])

    for rc in chybejici_pacienti:
        print 'Chybi BAL pacienta %s' % rc
    for rc, datum in chybejici_baly:
        print 'Chybi BAL pacienta %s pro %s' % (rc, datum)

def gen_summary(worksheet):
    for row in worksheet.iter_rows('A2:L8000'):
        try:
            rc = row[0].value.replace('/', '')
            jmeno = row[1].value
            datum_odberu = row[2].value.date()
            diagnoza = row[3].value.upper()
            if row[4].value == 1:
                alot = True
            elif row[4].value == 0:
                alot = False
            else:
                alot = None

            if row[10].value == 1:
                agranulocytoza = True
            elif row[10].value == 0:
                agranulocytoza = False
            else:
                agranulocytoza = None

            if row[11].value in ['EX', 'ano', 1]:
                exitus = True
            else:
                exitus = False
        except AttributeError:
            break

        yield rc, datum_odberu, {'diagnoza': diagnoza,
                                 'jmeno': jmeno,
                                 'alot': alot,
                                 'agranulocytoza': agranulocytoza,
                                 'exitus': exitus}

def gen_summary_2008style(worksheet):
    for row in worksheet.iter_rows('A5:AY8000'):
        try:
            rc = row[1].value.replace('/', '')
            jmeno = row[2].value
            datum_odberu = row[4].value.date()
            diagnoza = row[3].value
            if row[40].value == 1:
                alot = True
            elif row[40].value:
                alot = False
            else:
                alot = None

            if row[47].value == 1:
                agranulocytoza = True
            elif row[47].value:
                agranulocytoza = False
            else:
                agranulocytoza = None

            exitus = True if row[50].value.startswith('ano') else False
        except (AttributeError, TypeError):
            break

        yield rc, datum_odberu, {'diagnoza': diagnoza,
                                 'jmeno': jmeno,
                                 'alot': alot,
                                 'agranulocytoza': agranulocytoza,
                                 'exitus': exitus}

def gen_bal(worksheet):
    for row in worksheet.iter_rows('B2:H8000'):
        try:
            pp = row[0].value or ''
            vysledek = row[1].value or ''
            nalez = row[2].value
            memo = row[3].value or ''
            rc = row[4].value.replace('/', '')
            jmeno = row[5].value
            datum_odberu = row[6].value.date()
#             if rc == '485909248':
#                 print rc, row[6].value, datum_odberu

        except AttributeError:
            break

        yield rc, datum_odberu, {'pp': pp,
                                 'jmeno': jmeno,
                                 'vysledek': vysledek,
                                 'nalez': nalez,
                                 'memo': memo}

def process_2008style_xlsx(filez_list, bal_file):
    for f in filez_list:
        print 'Processing %s ...' % f
        wb = load_workbook(f)
        import_summary(gen_summary_2008style(wb.worksheets[0]))

    wb = load_workbook(bal_file)
    print 'Processing 2008'
    import_bal(gen_bal(wb.worksheets[0]))
    print 'Processing 2009'
    import_bal(gen_bal(wb.worksheets[1]))
    print 'Processing 2010'
    import_bal(gen_bal(wb.worksheets[2]))


def process_xlsx(fileobj):
    wb = load_workbook(fileobj)
    import_summary(gen_summary(wb.worksheets[0]))
    import_bal(gen_bal(wb.worksheets[1]))

if __name__ == '__main__':
    filez = glob.glob(os.path.join(sys.argv[1], 'BAL 20??.xlsx'))
    process_2008style_xlsx(filez, os.path.join(sys.argv[1], 'BAL pro OKH 2008-10_20150302_0822.xlsx'))

    filez = glob.glob(os.path.join(sys.argv[1], 'BAL pro OKH 20??_*.xlsx'))
    for f in filez:
        print 'Processing %s ...' % f
        process_xlsx(f)
